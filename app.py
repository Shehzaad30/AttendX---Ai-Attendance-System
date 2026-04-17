from flask import Flask, Response, render_template, redirect, request, url_for, session,jsonify, flash
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
from io import BytesIO

import os
import math
import uuid
import face_recognition
from fpdf import FPDF
import base64
from flask import render_template_string, make_response
import numpy as np
import cv2
from datetime import datetime, timedelta
import qrcode


import pymysql
app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024  
app.secret_key = "attendx_secret_key"
conn=pymysql.connect(host='localhost', user='root', password='', db='db_atten')
UPLOAD_FOLDER = 'static/student_img_upload/'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def get_admin_analytics_data():
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM student_data")
    total_students = cursor.fetchone()[0] or 0

    cursor.execute("""
        SELECT COUNT(*)
        FROM attendance
        WHERE date = CURDATE() AND status = 'Present'
    """)
    active_presence = cursor.fetchone()[0] or 0

    absent_today = max(total_students - active_presence, 0)
    stability_index = round((active_presence / total_students) * 100, 1) if total_students else 0

    labels = []
    values = []

    cursor.execute("""
        SELECT DATE(date) as day,
        COUNT(*) as total,
        SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) as present
        FROM attendance
        WHERE date >= CURDATE() - INTERVAL 6 DAY
        GROUP BY DATE(date)
        ORDER BY DATE(date)
    """)

    rows = cursor.fetchall()

    for row in rows:
        labels.append(row[0].strftime('%d %b'))  # e.g. 12 Apr
        total = row[1]
        present = row[2] or 0
        values.append(round((present / total) * 100) if total else 0)

    cursor.close()

    return {
        "labels": labels,
        "values": values,
        "total_students": total_students,
        "active_presence": active_presence,
        "absent_today": absent_today,
        "stability_index": stability_index,
        "pie_labels": ["Present", "Absent"],
        "pie_values": [active_presence, absent_today]
    }

@app.route("/alogin_process", methods=["POST"])
def alogin_process():
    admin_email = request.form.get('admin_email')
    admin_pass = request.form.get('admin_pass')
    cursor = conn.cursor() 
    cursor.execute("SELECT * FROM admin_login WHERE admin_email=%s", (admin_email,))
    admin = cursor.fetchone()
    if admin:
        if check_password_hash(admin[3], admin_pass):
            session['admin_id'] = admin[0]
            session['admin_name'] = admin[1]
            session['admin_email'] = admin[2]
            return redirect(url_for('dashboard'))
    return render_template("Admin/admin_login.html", error="Invalid email or password")

@app.route("/dashboard")
def dashboard():

    if 'admin_id' not in session:
        return redirect(url_for('login'))
    cursor = conn.cursor()
    query="SELECT COUNT(*) FROM student_data"
    cursor.execute(query)
    studata=cursor.fetchone()[0]
    query1 = "SELECT COUNT(*) FROM department_data"
    cursor.execute(query1)
    deptdata = cursor.fetchone()[0]
    query2 = "SELECT COUNT(*) FROM attendance"
    cursor.execute(query2)
    attenddata = cursor.fetchone()[0]
    query3 = "SELECT COUNT(*) FROM faculty_data"
    
    cursor.execute(query3)
    facultydata = cursor.fetchone()[0]

    query4 = "SELECT * from class_data Join faculty_data on class_data.Faculty_id=faculty_data.Faculty_id Join department_data on class_data.Department_id=department_data.Department_id"
    cursor.execute(query4)
    classes = cursor.fetchall()
    classdata = len(classes)
    
    chart_data = {
        "labels": ["Students", "Faculty", "Classes"],
        "values": [studata, facultydata, classdata]
    }
    
    cursor.close()

    return render_template(
        "Admin/dashboard.html",
        studata=studata,
        deptdata=deptdata,
        attenddata=attenddata,
        facultydata=facultydata,
        classes=classes,
        analytics=get_admin_analytics_data(),
        chart_data=chart_data
    )


@app.route("/add_department")
def add_department():
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    return render_template("Admin/add_department.html")

@app.route("/insert_department", methods=["POST"])
def insert_department():
    department_name = request.form['Department_name'] 
    department_code = request.form['Department_code'] 
    hod_name = request.form['HOD_name']
    cursor = conn.cursor() 
    cursor.execute("INSERT INTO department_data (Department_name, Department_code, HOD_name) VALUES (%s, %s, %s)", (department_name, department_code, hod_name))
    conn.commit() 
    cursor.close() 
    return redirect(url_for('view_department'))

@app.route("/view_department")
def view_department():
    
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    cursor = conn.cursor()
    query = "SELECT *FROM department_data"
    cursor.execute(query)
    departments = cursor.fetchall()
    
    return render_template("Admin/view_department.html", departments=departments)

@app.route("/delete_department/<int:Department_id>")
def delete_department(Department_id):
    cursor = conn.cursor()
    query = "DELETE FROM department_data WHERE Department_id=%s"
    val = (Department_id,)
    cursor.execute(query, val)
    conn.commit()
    cursor.close()
    return redirect(url_for('view_department'))

@app.route("/edit_department/<int:Department_id>")
def edit_department(Department_id):
    cursor = conn.cursor()
    query = "SELECT * FROM department_data WHERE Department_id=%s"
    cursor.execute(query, (Department_id,))
    department = cursor.fetchone()
    cursor.close()
    return render_template("Admin/edit_department.html", department=department)

@app.route("/edit_department_process/<int:Department_id>", methods=["POST"])
def edit_department_process(Department_id):
    department_name = request.form['Department_name']
    department_code = request.form['Department_code']
    hod_name = request.form['HOD_name']
    cursor = conn.cursor()
    query = "UPDATE department_data SET Department_name=%s, Department_code=%s, HOD_name=%s WHERE Department_id=%s"
    val = (department_name, department_code, hod_name, Department_id)
    cursor.execute(query, val)
    conn.commit()
    cursor.close()
    return redirect(url_for('view_department'))


@app.route("/")
def login():
    return render_template("Admin/admin_login.html")

@app.route("/add_user")
def add_user():
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    cursor = conn.cursor()


    query = "SELECT *FROM department_data"
    cursor.execute(query)
    departments = cursor.fetchall()

    return render_template("Admin/add_user.html",departments=departments)

@app.route("/api/analytics")
def api_analytics():
    if 'admin_id' not in session:
        return jsonify({"error": "Unauthorized"}), 401

    return jsonify(get_admin_analytics_data())


@app.route("/studentprocess", methods=["POST", "GET"])
def studentprocess():

    if 'admin_id' not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401

    try:

        Student_name = request.form.get('student_name')
        Division = request.form.get('Division')
        Enrollment_no = request.form.get('Enrollment_no')
        Department_id = request.form.get('Department_id')
        Semester = request.form.get('Semester')
        Email = request.form.get('Email')
        password = request.form.get('password')
        contact = request.form.get('contact')
        password = generate_password_hash(password)

        filename = None

        if 'captured_image' in request.files:

            file = request.files['captured_image']

            if file and file.filename != "":

                unique_id = str(uuid.uuid4())
                extension = os.path.splitext(file.filename)[1].lower()

                allowed_extensions = ['.png', '.jpg', '.jpeg']

                if extension not in allowed_extensions:
                    return jsonify({"success": False, "message": "Invalid image format"})

                filename = unique_id + extension

                filepath = os.path.join(UPLOAD_FOLDER, filename)

                os.makedirs(UPLOAD_FOLDER, exist_ok=True)

                file.save(filepath)
                db_path = f"static/student_img_upload/{filename}"

                # --- NEW FACE VALIDATION LOGIC ---
                try:
                    loaded_img = face_recognition.load_image_file(filepath)
                    encodings = face_recognition.face_encodings(loaded_img)

                    if len(encodings) == 0:
                        os.remove(filepath) # Clean up invalid image
                        return jsonify({"success": False, "message": "No face detected in the webcam capture! Please try again."})
                    
                    if len(encodings) > 1:
                        os.remove(filepath)
                        return jsonify({"success": False, "message": "Multiple faces detected! Please ensure only the student is in the frame."})
                        
                except Exception as eval_e:
                    print("Face validation error:", eval_e)
                    return jsonify({"success": False, "message": "Could not validate face image."})
                # ---------------------------------

        cursor = conn.cursor()

        query = """
        INSERT INTO student_data
        (Student_name, Division, Enrollment_no, Department_id,
        Semester, Email, password, contact, img_of_student)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """
        

        values = (
            Student_name,
            Division,
            Enrollment_no,
            Department_id,
            Semester,
            Email,
            password,
            contact,
            db_path     
                
        )

        cursor.execute(query, values)
        conn.commit()

        cursor.close()

        return jsonify({"success": True})

    except Exception as e:

        print("Error:", e)

        return jsonify({
            "success": False,
            "message": "Student insertion failed"
        })

@app.route("/view_user")
def view_user():

    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    cursor = conn.cursor()
    query = "SELECT * FROM student_data join department_data on student_data.Department_id=department_data.Department_id"
    cursor.execute(query,)
    students = cursor.fetchall()
    return render_template("Admin/view_user.html", students=students)

@app.route("/delete_user/<int:Student_id>")
def delete_user(Student_id):
    cursor = conn.cursor()
    query = "DELETE FROM student_data WHERE Student_id=%s"
    val = (Student_id,) 
    cursor.execute(query, val)
    conn.commit()
    cursor.close()
    return redirect(url_for('view_user'))

@app.route("/edit_user/<int:Student_id>")
def edit_user(Student_id):
    cursor = conn.cursor()


    query = "SELECT * FROM student_data WHERE Student_id=%s"
    cursor.execute(query, (Student_id,))
    student = cursor.fetchone()

    query = "SELECT * FROM department_data"
    cursor.execute(query)
    departments = cursor.fetchall()

    cursor.close()
    return render_template(
        "Admin/edit_user.html",
        student=student,
        departments=departments
    )

@app.route("/edit_user_process/<int:Student_id>", methods=["POST"])
def edit_user_process(Student_id):
    student_name = request.form['student_name']
    Division = request.form['Division']
    Enrollment_no = request.form['Enrollment_no']
    Department_id = request.form['Department_id']
    Semester = request.form['Semester']
    contact = request.form['contact']       
    cursor = conn.cursor()
    query = "UPDATE student_data SET Student_name=%s, Division=%s, Enrollment_no=%s, Department_id=%s, Semester=%s, Contact=%s WHERE Student_id=%s"
    val = (student_name, Division, Enrollment_no, Department_id, Semester, contact, Student_id)
    cursor.execute(query, val)
    conn.commit()
    cursor.close()
    return redirect(url_for('view_user',))


@app.route("/add_faculty")
def add_faculty():

    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    cursor = conn.cursor()
    query = "SELECT *FROM department_data"
    cursor.execute(query)
    departments = cursor.fetchall()
    cursor.close()
    return render_template("Admin/add_faculty.html",departments=departments)

@app.route("/insert_faculty", methods=["POST"])
def insert_faculty():
    Faculty_name = request.form['Faculty_name']
    Faculty_email = request.form['Faculty_email']
    Department_id = request.form['Department_id']
    contact = request.form['contact']
    Password = request.form['Password']
    password = generate_password_hash(Password)
    cursor = conn.cursor()
    
    select_query = "SELECT * FROM faculty_data WHERE Faculty_email=%s"
    cursor.execute(select_query, (Faculty_email,))
    existing_faculty = cursor.fetchone()
    if existing_faculty:
        cursor.close()
        return render_template("Admin/add_faculty.html", error="A faculty member with this email already exists.")
    
    query = "INSERT INTO faculty_data(Faculty_name, Faculty_email, Department_id, contact, Password) VALUES (%s, %s, %s, %s, %s)"
    cursor.execute(query, (Faculty_name, Faculty_email, Department_id, contact, Password))
    conn.commit()
    cursor.close()
    return redirect(url_for('view_faculty'))

@app.route("/view_faculty")
def view_faculty():
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    cursor = conn.cursor()
    query = "SELECT f.Faculty_id, f.Faculty_name, f.Faculty_email, f.Department_id, f.Contact, d.Department_name FROM faculty_data f JOIN department_data d on f.Department_id=d.Department_id"
    cursor.execute(query)
    faculty_members = cursor.fetchall()
    return render_template("Admin/view_faculty.html", faculty_members=faculty_members)

@app.route("/delete_faculty/<int:Faculty_id>")
def delete_faculty(Faculty_id):
    cursor = conn.cursor()
    query = "DELETE FROM faculty_data WHERE Faculty_id=%s"
    val = (Faculty_id,)
    cursor.execute(query, val)
    conn.commit()
    cursor.close()
    return redirect(url_for('view_faculty'))

@app.route("/edit_faculty/<int:Faculty_id>")
def edit_faculty(Faculty_id):
    cursor = conn.cursor()
    query = "SELECT * FROM faculty_data WHERE Faculty_id=%s"
    cursor.execute(query, (Faculty_id,))
    faculty = cursor.fetchone()

    query = "SELECT * FROM department_data"
    cursor.execute(query)
    departments = cursor.fetchall()

    cursor.close()
    return render_template(
        "Admin/edit_faculty.html",
        faculty=faculty,
        departments=departments
    )

@app.route("/edit_faculty_process/<int:Faculty_id>", methods=["POST"])
def edit_faculty_process(Faculty_id):
    Faculty_name = request.form['Faculty_name']
    Department_id = request.form['Department_id']
    contact = request.form['contact']

    cursor = conn.cursor()
    query = "UPDATE faculty_data SET Faculty_name=%s, Department_id=%s, contact=%s WHERE Faculty_id=%s"
    val = (Faculty_name, Department_id, contact, Faculty_id)
    cursor.execute(query, val)
    conn.commit()
    cursor.close()
    return redirect(url_for('view_faculty'))

@app.route("/add_class")
def add_class():
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    cursor = conn.cursor()
    query = "SELECT *FROM faculty_data"
    cursor.execute(query)
    faculty_members = cursor.fetchall()
    
    
    query = "SELECT *FROM department_data"
    cursor.execute(query)
    departments = cursor.fetchall()
    cursor.close()

    return render_template("Admin/add_class.html",faculty_members=faculty_members,departments=departments)

@app.route("/insert_class", methods=["POST"])
def insert_class():
    class_name = request.form['class_name']
    Faculty_id = request.form['Faculty_id']
    Department_id = request.form['Department_id']
    semester = request.form['semester']
    room_no = request.form['room_no']
    date = request.form['date']
    start_time = request.form['start_time']
    end_time = request.form['end_time']
    cursor = conn.cursor()
    query = "INSERT INTO class_data(class_name, Faculty_id, Department_id, semester, room_no, start_time, end_time) VALUES (%s, %s, %s, %s,%s,%s,%s)"
    cursor.execute(query, (class_name, Faculty_id, Department_id, semester, room_no, start_time, end_time))
    conn.commit()
    cursor.close()
    return redirect(url_for('view_class'))

@app.route("/view_class")
def view_class():

    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    cursor = conn.cursor()
    query = "SELECT * from class_data Join faculty_data on class_data.Faculty_id=faculty_data.Faculty_id Join department_data on class_data.Department_id=department_data.Department_id"
    cursor.execute(query)
    classes = cursor.fetchall()
    cursor.close()
    return render_template("Admin/view_class.html", classes=classes)

@app.route("/delete_class/<int:class_id>")
def delete_class(class_id):
    cursor = conn.cursor()
    query = "DELETE FROM class_data WHERE class_id=%s"
    val = (class_id,)
    cursor.execute(query, val)
    conn.commit()
    cursor.close()
    return redirect(url_for('view_class'))

@app.route("/edit_class/<int:class_id>")
def edit_class(class_id):
    cursor = conn.cursor()
    query = "SELECT * FROM class_data WHERE Class_id=%s"
    cursor.execute(query, (class_id,))
    class_data = cursor.fetchone()

    query = "SELECT * FROM faculty_data"
    cursor.execute(query)
    faculty_members = cursor.fetchall()

    query = "SELECT * FROM department_data"
    cursor.execute(query)
    departments = cursor.fetchall()

    cursor.close()
    return render_template(
        "Admin/edit_class.html",
        class_data = class_data,
        faculty_members=faculty_members,
        departments=departments
    )

@app.route("/edit_class_process/<int:class_id>", methods=["POST"])
def edit_class_process(class_id):
    class_name = request.form['class_name']
    Faculty_id = request.form['Faculty_id']
    Department_id = request.form['Department_id']
    semester = request.form['semester']
    room_no = request.form['room_no']
    date = request.form['date']
    start_time = request.form['start_time']
    end_time = request.form['end_time']

    cursor = conn.cursor()
    query = "UPDATE class_data SET class_name=%s, Faculty_id=%s, Department_id=%s, semester=%s, room_no=%s, date=%s, start_time=%s, end_time=%s WHERE class_id=%s"
    val = (class_name, Faculty_id, Department_id, semester, room_no, date, start_time, end_time, class_id)
    cursor.execute(query, val)
    conn.commit()
    cursor.close()
    return redirect(url_for('view_class'))

@app.route("/export_pdf/<int:Student_id>")
def export_pdf(Student_id):
    cursor = conn.cursor()

    query = """
    SELECT 
    s.Student_name,
    s.Enrollment_no,
    d.Department_name,
    s.Semester,
    a.date,
    a.status
    FROM attendance a
    JOIN student_data s ON a.Student_id = s.Student_id
    JOIN department_data d ON s.Department_id = d.Department_id
    WHERE a.Student_id = %s
    """

    cursor.execute(query, (Student_id,))
    data = cursor.fetchall()
    cursor.close()

    if not data:
        flash("No attendance records to export for this student.", "warning")
        return redirect(url_for('report'))

    # FPDF Generation
    pdf = FPDF()
    pdf.add_page()
    
    # Title
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(190, 10, txt="AttendX Attendance Report", ln=True, align='C')
    pdf.ln(10)

    # Student Info
    pdf.set_font("Arial", size=12)
    pdf.cell(190, 8, txt=f"Student Name: {data[0][0]}", ln=True)
    pdf.cell(190, 8, txt=f"Enrollment No: {data[0][1]}", ln=True)
    pdf.cell(190, 8, txt=f"Department: {data[0][2]}", ln=True)
    pdf.cell(190, 8, txt=f"Semester: {data[0][3]}", ln=True)
    pdf.ln(10)

    # Table Header
    pdf.set_font("Arial", 'B', 12)
    pdf.set_fill_color(240, 240, 240) # light gray summary header
    pdf.cell(95, 10, 'Date', 1, 0, 'C', fill=True)
    pdf.cell(95, 10, 'Status', 1, 1, 'C', fill=True)
    
    # Table Content
    pdf.set_font("Arial", '', 12)
    for row in data:
        pdf.cell(95, 10, str(row[4]), 1, 0, 'C')
        # Check status length or content
        status_text = str(row[5])
        pdf.cell(95, 10, status_text, 1, 1, 'C')

    # Output to response
    pdf_bytes = pdf.output(dest='S').encode('latin-1')

    # For modern download filename (spaces swapped to underscores)
    safe_name = str(data[0][0]).replace(' ', '_')
    
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={safe_name}_attendance.pdf'

    return response

@app.route("/export_all_pdf")
def export_all_pdf():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    cursor = conn.cursor()
    query = """
    SELECT 
        a.Student_name,
        a.Enrollment_no,
        b.Department_name,
        a.Semester
    FROM student_data a
    JOIN department_data b 
    ON a.Department_id = b.Department_id
    """
    cursor.execute(query)
    data = cursor.fetchall()
    cursor.close()

    if not data:
        flash("No student records found to export.", "warning")
        return redirect(url_for('report'))

    pdf = FPDF()
    pdf.add_page()
    
    # Title
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(190, 10, txt="AttendX Master Student Report", ln=True, align='C')
    pdf.ln(10)

    # Table Header
    pdf.set_font("Arial", 'B', 10)
    pdf.set_fill_color(240, 240, 240)
    # Widths: Name(55), Enrollment(45), Dept(60), Sem(30)
    pdf.cell(10, 10, '#', 1, 0, 'C', fill=True)
    pdf.cell(50, 10, 'Student Name', 1, 0, 'C', fill=True)
    pdf.cell(45, 10, 'Enrollment No', 1, 0, 'C', fill=True)
    pdf.cell(55, 10, 'Department', 1, 0, 'C', fill=True)
    pdf.cell(30, 10, 'Semester', 1, 1, 'C', fill=True)
    
    # Table Content
    pdf.set_font("Arial", '', 10)
    for idx, row in enumerate(data, start=1):
        pdf.cell(10, 10, str(idx), 1, 0, 'C')
        pdf.cell(50, 10, str(row[0]), 1, 0, 'C')
        pdf.cell(45, 10, str(row[1]), 1, 0, 'C')
        pdf.cell(55, 10, str(row[2]), 1, 0, 'C')
        pdf.cell(30, 10, str(row[3]), 1, 1, 'C')

    pdf_bytes = pdf.output(dest='S').encode('latin-1')
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=Master_Student_Report.pdf'

    return response

@app.route("/report")
def report():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    cursor = conn.cursor()

    query = """
    SELECT 
        a.Student_id,
        a.Student_name,
        a.Enrollment_no,
        a.Semester,
        b.Department_name
    FROM student_data a
    JOIN department_data b 
    ON a.Department_id = b.Department_id
    """

    cursor.execute(query)
    data = cursor.fetchall()
    cursor.close()

    return render_template("Admin/report.html", data=data)

@app.route("/viewattendance/<int:Student_id>")
def view_attendance(Student_id):
    cursor = conn.cursor()

    query = """
    SELECT 
    s.Student_name,
    s.Enrollment_no,
    d.Department_name,
    s.Semester,
    a.date,
    a.status
    FROM attendance a
    JOIN student_data s ON a.Student_id = s.Student_id
    JOIN department_data d ON s.Department_id = d.Department_id
    WHERE a.Student_id = %s
    """

    cursor.execute(query, (Student_id,))
    data = cursor.fetchall()

    cursor.close()

    return render_template("Admin/view_attendance.html", data=data)

@app.route("/attendance_rule")
def attendance():

    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    return render_template("Admin/attendance_rule.html")

@app.route("/allowed-time_window")
def settings():

    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    return render_template("Admin/allowed_time_window.html")

@app.route("/gps_range")
def gps_range():

    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    return render_template("Admin/gps_range.html")

@app.route("/camera_setting")
def camera_setting():

    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    return render_template("Admin/camera_setting.html")

@app.route("/backup_database")
def backup_database():

    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    return render_template("Admin/backup_database.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route("/faculty_login_process", methods=["POST"])
def faculty_login_process():
    Faculty_email = request.form['Faculty_email']
    Password = request.form['Password']
    cursor = conn.cursor()
    query = '''SELECT f.Faculty_id, f.Faculty_name, f.Faculty_email, f.Department_id, f.Contact, d.Department_name, f.Password FROM faculty_data f 
    JOIN department_data d on f.Department_id=d.Department_id WHERE Faculty_email=%s '''
    cursor.execute(query, (Faculty_email))
    faculty = cursor.fetchone()
    if faculty and check_password_hash(faculty[6], Password):
        session['faculty_id'] = faculty[0]
        session['Faculty_name'] = faculty[1]
        session['Faculty_email'] = faculty[2]
        session['Department_id'] = faculty[3]
        session['contact'] = faculty[4]
        session['Department_name'] = faculty[5]
        return redirect(url_for('faculty_dashboard'))
    else:
        flash("Invalid Email or Password")
        return redirect(url_for("faculty_login"))
    
@app.route("/faculty_login")
def faculty_login():
    return render_template("Faculty/faculty_login.html")

@app.route("/faculty_dashboard")
def faculty_dashboard():
    if 'faculty_id' not in session:
        return redirect(url_for('faculty_login'))

    faculty_id = session['faculty_id']
    department_id = session['Department_id']

    cursor = conn.cursor(pymysql.cursors.DictCursor)

    cursor.execute(
        "SELECT COUNT(*) AS total FROM class_data WHERE Faculty_id=%s",
        (faculty_id,)
    )
    classdata = cursor.fetchone()['total']

    cursor.execute(
        "SELECT COUNT(*) AS total FROM student_data WHERE Department_id=%s",
        (department_id,)
    )
    total_students = cursor.fetchone()['total']

    cursor.execute("""
        SELECT *
        FROM class_data
        WHERE Faculty_id=%s AND date = CURDATE()
        ORDER BY start_time
    """, (faculty_id,))
    today_classes = cursor.fetchall()

    today_classes_count = len(today_classes)

    cursor.execute("""
        SELECT COUNT(*) AS total
        FROM attendance
        WHERE date = CURDATE()
    """)
    today_attendance = cursor.fetchone()['total']

    pending_classes = today_classes_count - today_attendance
    if pending_classes < 0:
        pending_classes = 0

    cursor.execute("""
        SELECT *
        FROM class_data
        WHERE Faculty_id=%s AND date = CURDATE()
        ORDER BY start_time
        LIMIT 1
    """, (faculty_id,))
    next_class = cursor.fetchone()

    cursor.close()

    return render_template(
        "Faculty/faculty_dashboard.html",
        classdata=classdata,
        total_students=total_students,
        today_attendance=today_attendance,
        pending_classes=pending_classes,
        today_classes=today_classes,
        next_class=next_class
    )

@app.route("/faculty_layout")
def faculty_layout():
    if 'faculty_id' not in session:
        return redirect(url_for('faculty_login'))
    
    return render_template("Faculty/layout.html")

@app.route("/myclasses")
def myclasses():

    if 'faculty_id' not in session:
        return redirect(url_for('faculty_login'))
    faculty_id = session['faculty_id']
    query = "SELECT class_data.*, faculty_data.Faculty_name, department_data.Department_name FROM class_data " \
    "JOIN faculty_data ON class_data.Faculty_id = faculty_data.Faculty_id " \
    "JOIN department_data ON class_data.Department_id = department_data.Department_id WHERE class_data.Faculty_id = %s"
    cursor = conn.cursor()
    cursor.execute(query, (faculty_id,))
    classes = cursor.fetchall()
    cursor.close()
    
    return render_template("Faculty/myclasses.html", classes=classes)

@app.route("/start_attendance")
def show_attendance():

    if 'faculty_id' not in session:
        return redirect(url_for('faculty_login'))
    
    return render_template("Faculty/start_attendance.html")

@app.route("/faculty_student_report")
def faculty_student_report():

    if 'faculty_id' not in session:
        return redirect(url_for('faculty_login'))
    cursor = conn.cursor()

    query = """
    SELECT 
        a.Student_id,
        a.Student_name,
        a.Enrollment_no,
        a.Semester,
        b.Department_name
    FROM student_data a
    JOIN department_data b 
    ON a.Department_id = b.Department_id
    """

    cursor.execute(query)
    data2 = cursor.fetchall()
    cursor.close()

    return render_template("Faculty/faculty_student_report.html", data2=data2)

@app.route("/viewattendance_student/<int:Student_id>")
def viewattendance_student(Student_id):
    cursor = conn.cursor()

    query = """
    SELECT 
    s.Student_name,
    s.Enrollment_no,
    d.Department_name,
    s.Semester,
    a.date,
    a.status
    FROM attendance a
    JOIN student_data s ON a.Student_id = s.Student_id
    JOIN department_data d ON s.Department_id = d.Department_id
    WHERE a.Student_id = %s
    """

    cursor.execute(query, (Student_id,))
    data = cursor.fetchall()

    cursor.close()

    return render_template("Faculty/viewattendance_student.html", data=data)

@app.route("/export_pdf_faculty/<int:Student_id>")
def export_pdf_faculty(Student_id):
    cursor = conn.cursor()

    query = """
    SELECT 
    s.Student_name,
    s.Enrollment_no,
    d.Department_name,
    s.Semester,
    a.date,
    a.status
    FROM attendance a
    JOIN student_data s ON a.Student_id = s.Student_id
    JOIN department_data d ON s.Department_id = d.Department_id
    WHERE a.Student_id = %s
    """

    cursor.execute(query, (Student_id,))
    data = cursor.fetchall()
    cursor.close()

    if not data:
        flash("No attendance records to export for this student.", "warning")
        return redirect(url_for('faculty_student_report'))

    # FPDF Generation
    pdf = FPDF()
    pdf.add_page()
    
    # Title
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(190, 10, txt="AttendX Attendance Report", ln=True, align='C')
    pdf.ln(10)

    # Student Info
    pdf.set_font("Arial", size=12)
    pdf.cell(190, 8, txt=f"Student Name: {data[0][0]}", ln=True)
    pdf.cell(190, 8, txt=f"Enrollment No: {data[0][1]}", ln=True)
    pdf.cell(190, 8, txt=f"Department: {data[0][2]}", ln=True)
    pdf.cell(190, 8, txt=f"Semester: {data[0][3]}", ln=True)
    pdf.ln(10)

    # Table Header
    pdf.set_font("Arial", 'B', 12)
    pdf.set_fill_color(240, 240, 240) # light gray summary header
    pdf.cell(95, 10, 'Date', 1, 0, 'C', fill=True)
    pdf.cell(95, 10, 'Status', 1, 1, 'C', fill=True)
    
    # Table Content
    pdf.set_font("Arial", '', 12)
    for row in data:
        pdf.cell(95, 10, str(row[4]), 1, 0, 'C')
        # Check status length or content
        status_text = str(row[5])
        pdf.cell(95, 10, status_text, 1, 1, 'C')

    # Output to response
    pdf_bytes = pdf.output(dest='S').encode('latin-1')

    # For modern download filename (spaces swapped to underscores)
    safe_name = str(data[0][0]).replace(' ', '_')
    
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={safe_name}_attendance.pdf'

    return response

@app.route("/export_all_pdf_faculty")
def export_all_pdf_faculty():
    if 'faculty_id' not in session:
        print("Unauthorized access attempt to export_all_pdf")
        return redirect(url_for('faculty_login'))

    cursor = conn.cursor()
    query = """
    SELECT 
        a.Student_name,
        a.Enrollment_no,
        b.Department_name,
        a.Semester
    FROM student_data a
    JOIN department_data b 
    ON a.Department_id = b.Department_id
    """
    cursor.execute(query)
    data = cursor.fetchall()
    cursor.close()
    print("something")
    if not data:
        flash("No student records found to export.", "warning")
        return redirect(url_for('faculty_student_report'))

    pdf = FPDF()
    pdf.add_page()
    
    # Title
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(190, 10, txt="AttendX Master Student Report", ln=True, align='C')
    pdf.ln(10)

    # Table Header
    pdf.set_font("Arial", 'B', 10)
    pdf.set_fill_color(240, 240, 240)
    # Widths: Name(55), Enrollment(45), Dept(60), Sem(30)
    pdf.cell(10, 10, '#', 1, 0, 'C', fill=True)
    pdf.cell(50, 10, 'Student Name', 1, 0, 'C', fill=True)
    pdf.cell(45, 10, 'Enrollment No', 1, 0, 'C', fill=True)
    pdf.cell(55, 10, 'Department', 1, 0, 'C', fill=True)
    pdf.cell(30, 10, 'Semester', 1, 1, 'C', fill=True)
    
    # Table Content
    pdf.set_font("Arial", '', 10)
    for idx, row in enumerate(data, start=1):
        pdf.cell(10, 10, str(idx), 1, 0, 'C')
        pdf.cell(50, 10, str(row[0]), 1, 0, 'C')
        pdf.cell(45, 10, str(row[1]), 1, 0, 'C')
        pdf.cell(55, 10, str(row[2]), 1, 0, 'C')
        pdf.cell(30, 10, str(row[3]), 1, 1, 'C')

    pdf_bytes = pdf.output(dest='S').encode('latin-1')
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=Master_Student_Report.pdf'

    return response

@app.route("/faculty_profile")
def faculty_profile():
    if 'faculty_id' not in session:
        return redirect(url_for('faculty_login'))
    
    return render_template("Faculty/faculty_profile.html")

@app.route("/faculty_edit_profile")
def faculty_edit_profile():
    if 'faculty_id' not in session:
        return redirect(url_for('faculty_login'))
    
    cursor = conn.cursor()
    faculty_id = session['faculty_id']

    cursor.execute("""
        SELECT Faculty_id, Faculty_name, Department_id, contact
        FROM faculty_data
        WHERE Faculty_id=%s
    """, (faculty_id,))

    faculty = cursor.fetchall()

    cursor.execute("SELECT * FROM department_data")
    departments = cursor.fetchall()

    cursor.close()

    return render_template(
        "Faculty/faculty_edit_profile.html",
        faculty=faculty,
        departments=departments
    )

@app.route("/edit_faculty_profile_process/<int:Faculty_id>", methods=["POST"])
def edit_faculty_profile_process(Faculty_id):
    Faculty_name = request.form['Faculty_name']
    Department_id = request.form['Department_id']
    contact = request.form['contact']

    cursor = conn.cursor()
    query = "UPDATE faculty_data SET Faculty_name=%s, Department_id=%s, contact=%s WHERE Faculty_id=%s"
    val = (Faculty_name, Department_id, contact, Faculty_id)
    cursor.execute(query, val)
    conn.commit()
    cursor.close()
    return redirect(url_for('faculty_profile'))

@app.route("/faculty_logout")
def faculty_logout():
    session.clear()
    return redirect(url_for('faculty_login'))

@app.route("/student_login_process", methods=["POST"])
def student_login_process():
    Enrollment_no = request.form['Enrollment_no']
    password = request.form['password']

    cursor = conn.cursor()

    query = """
        SELECT s.Student_id, s.Student_name, s.Enrollment_no,
            s.Department_id, s.Email, s.contact,
            d.Department_name, s.password
        FROM student_data s
        JOIN department_data d
        ON s.Department_id = d.Department_id
        WHERE s.Enrollment_no=%s
    """

    cursor.execute(query, (Enrollment_no,))
    student = cursor.fetchone()
    conn.commit()
    cursor.close()
    print(student[7])

    if student and check_password_hash(student[7], password):
        session['Student_id'] = student[0]
        session['Student_name'] = student[1]
        session['Enrollment_no'] = student[2]
        session['Department_id'] = student[3]
        session['Email'] = student[4]
        session['contact'] = student[5]
        session['Department_name'] = student[6]

        return redirect(url_for('student_dashboard'))
    else:
        flash("Invalid Enrollment No or Password")
        return redirect(url_for('student_login'))



@app.route("/student_login")
def student_login():

    return render_template("Student/student_login.html")

@app.route("/student_dashboard")
def student_dashboard():

    if 'Student_id' not in session:
        return redirect(url_for('student_login'))

    student_id = session['Student_id']
    department_id = session['Department_id']

    cursor = conn.cursor(pymysql.cursors.DictCursor)

    cursor.execute(
        "SELECT COUNT(*) as total FROM class_data WHERE Department_id=%s",
        (department_id,))
    total_classes = cursor.fetchone()['total']

    cursor.execute("""
    SELECT *
    FROM class_data
    WHERE Department_id=%s AND date=CURDATE()
    """,(department_id,))
    today_classes = cursor.fetchall()

    today_classes_count = len(today_classes)

    cursor.execute("""
        SELECT 
            COUNT(*) as total_classes,
            SUM(CASE WHEN status = 'Present' THEN 1 ELSE 0 END) as present_count
        FROM attendance
        WHERE Student_id=%s
        """, (student_id,))

    data = cursor.fetchone()

    total_classes = data['total_classes']
    present_count = data['present_count'] if data['present_count'] else 0

    attendance_percentage = 0
    if total_classes > 0:
        attendance_percentage = round((present_count / total_classes) * 100)

    cursor.execute("""
    SELECT *
    FROM class_data
    WHERE Department_id=%s AND date=CURDATE()
    ORDER BY start_time
    LIMIT 1
    """,(department_id,))
    next_class = cursor.fetchone()

    cursor.execute("""
    SELECT date, 
           SUM(CASE WHEN status = 'Present' THEN 1 ELSE 0 END) as present_count
    FROM attendance
    WHERE Student_id=%s
    GROUP BY date
    """, (student_id,))
    heatmap_rows = cursor.fetchall()
    
    heatmap_data = []
    for row in heatmap_rows:
        if row['date']:
            # Handle date string conversion
            date_str = row['date'].strftime('%Y-%m-%d') if hasattr(row['date'], 'strftime') else str(row['date'])
            heatmap_data.append([date_str, int(row['present_count'])])
            
    cursor.close()
    return render_template(
        "Student/student_dashboard.html",
        total_classes=total_classes,
        today_classes=today_classes,
        today_classes_count=today_classes_count,
        attendance_percentage=attendance_percentage,
        next_class=next_class,
        heatmap_data=heatmap_data
    )

@app.route("/classes")

def classes():

    if 'Student_id' not in session:
        return redirect(url_for('student_login'))
    student_id = session['Student_id']
    cursor = conn.cursor()


    cursor.execute(
        "SELECT Department_id FROM student_data WHERE Student_id=%s",
        (student_id,)

    )
    student = cursor.fetchone()
    if not student:
        cursor.close()
        return "Student not found"
    Department_id = student[0]
    cursor.execute(
        """SELECT c.class_id, c.class_name, f.faculty_name,d.department_name, c.room_no, c.date, c.start_time, c.end_time, c.latitude, c.longitude
        FROM class_data c
        JOIN faculty_data f ON c.faculty_id = f.faculty_id
        JOIN Department_data d ON c.Department_id=d.Department_id 
        WHERE c.Department_id=%s""",
        (Department_id,)
    )
    classes = cursor.fetchall()
    cursor.close()
    return render_template("Student/classes.html", classes = classes)

@app.route("/student_profile")
def student_profile():
    
    if 'Student_id' not in session:
        return redirect(url_for('student_login'))
    cursor = conn.cursor()
    student_id = session['Student_id']
    query="select * from student_data where Student_id=%s"
    cursor.execute(query, (student_id,))
    student = cursor.fetchall()
    cursor.close()
    
    return render_template("Student/student_profile.html", student=student)

@app.route("/student_edit_profile/<int:Student_id>")
def student_edit_profile(Student_id):
    
    if 'Student_id' not in session:
        return redirect(url_for('student_login'))
    
    cursor = conn.cursor()
    cursor.execute("""
        SELECT Student_id, Student_name, Department_id, contact
        FROM student_data
        WHERE Student_id=%s
    """, (Student_id,))

    student = cursor.fetchall()

    cursor.execute("SELECT * FROM department_data")
    departments = cursor.fetchall()

    cursor.close()

    return render_template("Student/edit_profile.html",
                        departments=departments,
                        student=student)

@app.route("/student_edit_profile_process", methods=['POST'] )
def student_edit_profile_process():
    Student_id = request.form['Student_id']

    Student_name = request.form['Student_name']
    Department_id = request.form['Department_id']
    contact = request.form['contact']

    cursor = conn.cursor()
    query = "UPDATE student_data SET Student_name=%s, Department_id=%s, contact=%s WHERE Student_id=%s"
    val = (Student_name, Department_id, contact, Student_id)
    cursor.execute(query, val)
    conn.commit()
    cursor.close()
    return redirect(url_for('student_profile'))

@app.route("/student_attendance_report")
def student_attendance_report():


    student_id = session['Student_id']

    cursor = conn.cursor()

    query = """
    SELECT 
        s.Student_name,
        s.Enrollment_no,
        d.Department_name,
        s.Semester,
        a.date,
        a.status
    FROM attendance a
    JOIN student_data s ON a.Student_id = s.Student_id
    JOIN department_data d ON s.Department_id = d.Department_id
    WHERE a.Student_id = %s
    """

    cursor.execute(query, (student_id,))
    data1 = cursor.fetchall()

    cursor.close()

    

    return render_template(
        "Student/student_attendance_report.html",
        data1=data1,
    
    )

@app.route("/export_pdf_student/<int:Student_id>")
def export_pdf_student(Student_id):
    cursor = conn.cursor()
    print("Exporting PDF for Student ID:", Student_id)
    query = """
    SELECT 
    s.Student_name,
    s.Enrollment_no,
    d.Department_name,
    s.Semester,
    a.date,
    a.status
    FROM attendance a
    JOIN student_data s ON a.Student_id = s.Student_id
    JOIN department_data d ON s.Department_id = d.Department_id
    WHERE a.Student_id = %s

    """

    cursor.execute(query, (Student_id,))
    data = cursor.fetchall()
    cursor.close()

    if not data:
        flash("No attendance records to export for this student.", "warning")
        return redirect(url_for('student_attendance_report'))

    # FPDF Generation
    pdf = FPDF()
    pdf.add_page()
    
    # Title
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(190, 10, txt="AttendX Attendance Report", ln=True, align='C')
    pdf.ln(10)

    # Student Info
    pdf.set_font("Arial", size=12)
    pdf.cell(190, 8, txt=f"Student Name: {data[0][0]}", ln=True)
    pdf.cell(190, 8, txt=f"Enrollment No: {data[0][1]}", ln=True)
    pdf.cell(190, 8, txt=f"Department: {data[0][2]}", ln=True)
    pdf.cell(190, 8, txt=f"Semester: {data[0][3]}", ln=True)
    pdf.ln(10)

    # Table Header
    pdf.set_font("Arial", 'B', 12)
    pdf.set_fill_color(240, 240, 240) # light gray summary header
    pdf.cell(95, 10, 'Date', 1, 0, 'C', fill=True)
    pdf.cell(95, 10, 'Status', 1, 1, 'C', fill=True)
    
    # Table Content
    pdf.set_font("Arial", '', 12)
    for row in data:
        pdf.cell(95, 10, str(row[4]), 1, 0, 'C')
        # Check status length or content
        status_text = str(row[5])
        pdf.cell(95, 10, status_text, 1, 1, 'C')

    # Output to response
    pdf_bytes = pdf.output(dest='S').encode('latin-1')

    # For modern download filename (spaces swapped to underscores)
    safe_name = str(data[0][0]).replace(' ', '_')
    
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={safe_name}_attendance.pdf'

    return response

@app.route("/export_excel_student/<int:Student_id>")
def export_excel_student(Student_id):

    if 'Student_id' not in session:
        return redirect(url_for('student_login'))

    cursor = conn.cursor(pymysql.cursors.DictCursor)

    query = """
    SELECT s.Student_name, s.Enrollment_no, d.Department_name,
        s.Semester, a.date, a.status
    FROM attendance a
    JOIN student_data s ON a.Student_id = s.Student_id
    JOIN department_data d ON s.Department_id = d.Department_id
    WHERE a.Student_id = %s
    ORDER BY a.date DESC
    """

    cursor.execute(query, (Student_id,))
    data = cursor.fetchall()
    cursor.close()

    # Create Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Headers
    headers = ["Student Name", "Enrollment No", "Department", "Semester", "Date", "Status"]
    ws.append(headers)

    # Data rows
    for row in data:
        ws.append([
            row['Student_name'],
            row['Enrollment_no'],
            row['Department_name'],
            row['Semester'],
            str(row['date']),
            row['status']
        ])

    # Styling (optional but good)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2

    # Save to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=attendance_report_{Student_id}.xlsx"
        }
    )


@app.route("/mark_attendance/<int:class_id>")
def mark_attendance(class_id):

    if 'Student_id' not in session:
        return redirect(url_for('student_login'))

    return render_template(
        "Student/mark_attendance.html",
        class_id=class_id
    )

# @app.route("/mark_attendanceprocess", methods=["POST"])
# def mark_attendanceprocess():


#     class_id = request.form.get('class_id')
#     student_id = session.get('Student_id')
#     if not student_id:
#         return redirect(url_for('student_login'))
#     if not class_id:
#         return "Invalid class ID"


#     image_data = request.form['image_data']
#     class_id = int(class_id)
#     cursor = conn.cursor(pymysql.cursors.DictCursor)
    

#     image_data = image_data.split(",")[1]
#     image_bytes = base64.b64decode(image_data)
#     np_arr = np.frombuffer(image_bytes, np.uint8)
#     frame = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)

#     rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

#     unknown_encodings = face_recognition.face_encodings(rgb_frame)

#     if len(unknown_encodings) == 0:
#         msg3="No face detected in the image. Please try again."
#         return render_template("Student/student_dashboard.html", msg3=msg3)
#         #return redirect(url_for('mark_attendance', class_id=class_id))

#     unknown_encoding = unknown_encodings[0]
#     if len(unknown_encodings) > 1:
#         msg4="Multiple faces detected! Please ensure only you are in the frame."
#         return render_template("Student/student_dashboard.html", msg4=msg4)
#         #return redirect(url_for('mark_attendance', class_id=class_id))

#     cursor.execute(
#         "SELECT Student_id, img_of_student FROM student_data WHERE Student_id=%s",
#         (student_id,)
#     )

#     student = cursor.fetchone()

#     if not student:
#         msg5="Student record not found in database."
#         return render_template("Student/student_dashboard.html", msg5=msg5)

#     # img_filename = student['img_of_student']
#     # img_path = os.path.join("static/student_img_upload", img_filename)
#     img_filename = student['img_of_student']
#     img_path = os.path.join("static/student_img_upload", img_filename)

#     if not os.path.exists(img_path):
#         msg6="Your registration image is missing. Please contact admin."
#         return render_template("Student/student_dashboard.html", msg6=msg6)

#     known_image = face_recognition.load_image_file(img_path)
#     known_encodings = face_recognition.face_encodings(known_image)

#     if len(known_encodings) == 0:
#         msg7="No face was found in your stored registration image."
#         return render_template("Student/student_dashboard.html", msg7=msg7)
#         return redirect(url_for('student_dashboard'))

#     known_encoding = known_encodings[0]


#     distance = face_recognition.face_distance([known_encoding], unknown_encoding)
#     # Relaxed tolerance from 0.5 to 0.55 to reduce false rejections due to lighting
#     if distance[0] < 0.55:
#         today = datetime.now().date()

#         cursor.execute("""
#         SELECT *
#         FROM attendance
#         WHERE Student_id=%s AND date=%s AND class_id=%s
#         """, (student['Student_id'], today, class_id))

#         already = cursor.fetchone()

#         if already:
#             #flash("Attendance has already been marked for this class today.", "warning")
#             msg='Attendance has already been marked for this class today.'
#             #return redirect(url_for('student_dashboard',msg=msg))
#             return render_template("Student/student_dashboard.html", msg=msg)

#         cursor.execute("""
#         INSERT INTO attendance (Student_id, class_id, date, status)
#         VALUES (%s, %s, %s, %s)
#         """, (student['Student_id'], class_id, today, "Present"))
#         conn.commit()
#         print("hi")
#         #flash("Attendance Marked Successfully!", "success")
#         msg1='Attendance Marked Successfully!'
#         return render_template("Student/student_dashboard.html", msg1=msg1)
    
#     msg2="Face does not match the registered user. Attendance denied."
#     return render_template("Student/student_dashboard.html", msg2=msg2)

#     return redirect(url_for('mark_attendance', class_id=class_id))



@app.route("/mark_attendanceprocess", methods=["POST"])
def mark_attendanceprocess():

    class_id = request.form.get('class_id')
    student_id = session.get('Student_id')

    if not student_id:
        return redirect(url_for('student_login'))

    if not class_id:
        return "Invalid class ID"

    class_id = int(class_id)
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    try:
        # ---------- CAPTURE IMAGE ----------
        image_data = request.form['image_data']
        image_data = image_data.split(",")[1]
        image_bytes = base64.b64decode(image_data)
        np_arr = np.frombuffer(image_bytes, np.uint8)
        frame = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)

        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

        # Improve accuracy
        rgb_frame = cv2.resize(rgb_frame, (0, 0), fx=0.5, fy=0.5)

        unknown_encodings = face_recognition.face_encodings(rgb_frame)

        if len(unknown_encodings) == 0:
            msg3 = "No face detected in the image. Please try again."
            return render_template("Student/student_dashboard.html", msg3=msg3)

        if len(unknown_encodings) > 1:
            msg4 = "Multiple faces detected! Please ensure only you are in the frame."
            return render_template("Student/student_dashboard.html", msg4=msg4)

        unknown_encoding = unknown_encodings[0]

        # ---------- FETCH STUDENT ----------
        cursor.execute(
            "SELECT Student_id, img_of_student FROM student_data WHERE Student_id=%s",
            (student_id,)
        )
        student = cursor.fetchone()

        if not student:
            msg5 = "Student record not found in database."
            return render_template("Student/student_dashboard.html", msg5=msg5)

        # ---------- FIX IMAGE PATH ----------
        img_path = student['img_of_student']

        # Handle both cases (filename or full path)
        if not img_path.startswith("static"):
            img_path = os.path.join("static/student_img_upload", img_path)

        print("Image Path Used:", img_path)

        if not os.path.exists(img_path):
            msg6 = "Your registration image is missing. Please contact admin."
            return render_template("Student/student_dashboard.html", msg6=msg6)

        # ---------- LOAD STORED IMAGE ----------
        known_image = face_recognition.load_image_file(img_path)
        known_encodings = face_recognition.face_encodings(known_image)

        if len(known_encodings) == 0:
            msg7 = "No face found in your stored image."
            return render_template("Student/student_dashboard.html", msg7=msg7)

        known_encoding = known_encodings[0]

        # ---------- FACE MATCH ----------
        matches = face_recognition.compare_faces([known_encoding], unknown_encoding, tolerance=0.55)
        distance = face_recognition.face_distance([known_encoding], unknown_encoding)

        print("Match:", matches[0])
        print("Distance:", distance[0])

        if matches[0]:

            today = datetime.now().date()

            cursor.execute("""
                SELECT * FROM attendance
                WHERE Student_id=%s AND date=%s AND class_id=%s
            """, (student['Student_id'], today, class_id))

            already = cursor.fetchone()

            if already:
                msg = "Attendance already marked for this class today."
                return render_template("Student/student_dashboard.html", msg=msg)

            cursor.execute("""
                INSERT INTO attendance (Student_id, class_id, date, status)
                VALUES (%s, %s, %s, %s)
            """, (student['Student_id'], class_id, today, "Present"))

            conn.commit()

            msg1 = "Attendance Marked Successfully!"
            return render_template("Student/student_dashboard.html", msg1=msg1)

        else:
            msg2 = "Face does not match the registered user."
            return render_template("Student/student_dashboard.html", msg2=msg2)

    except Exception as e:
        print("Error:", e)
        return render_template("Student/student_dashboard.html", msg="Error occurred")

@app.route('/save_location', methods=['POST','GET'])
def save_location():
    data = request.get_json()

    latitude = data.get('latitude')
    longitude = data.get('longitude')

    print("Latitude:", latitude)
    print("Longitude:", longitude)


    return latitude,longitude

@app.route("/student_logout")
def student_logout():
    session.clear()
    return redirect(url_for('student_login'))


# ================================================================
#  ATTENDANCE SYSTEM — Session Control, Initialization & Marking
# ================================================================

def get_db_cursor(dict_cursor=False):
    """Return a fresh cursor, reconnecting if needed."""
    conn.ping(reconnect=True)
    if dict_cursor:
        return conn.cursor(pymysql.cursors.DictCursor)
    return conn.cursor()


def is_session_active(class_id):
    """
    Return True if the current time falls within the class's
    start_time and end_time for today's date.
    """
    cursor = get_db_cursor(dict_cursor=True)
    cursor.execute(
        "SELECT start_time, end_time, date FROM class_data WHERE class_id = %s",
        (class_id,)
    )
    cls = cursor.fetchone()
    cursor.close()

    if not cls:
        return False

    now = datetime.now()

    # class_date check: only active on its scheduled date
    class_date = cls['date']
    if class_date and class_date != now.date():
        return False

    # start_time / end_time may come back as timedelta (pymysql quirk)
    def to_time(val):
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.time()
        # timedelta → seconds
        import datetime as dt
        if isinstance(val, dt.timedelta):
            total = int(val.total_seconds())
            h, rem = divmod(total, 3600)
            m, s = divmod(rem, 60)
            return dt.time(h, m, s)
        return val  # already time

    start = to_time(cls['start_time'])
    end   = to_time(cls['end_time'])

    if start is None or end is None:
        return True   # no time restriction configured

    current = now.time()
    return start <= current <= end


def is_late_for_class(class_id, tolerance_minutes=15):
    """
    Check if the current time is more than tolerance_minutes 
    past the class start_time.
    """
    cursor = get_db_cursor(dict_cursor=True)
    cursor.execute(
        "SELECT start_time FROM class_data WHERE class_id = %s AND date = CURDATE()",
        (class_id,)
    )
    cls = cursor.fetchone()
    cursor.close()

    if not cls or not cls['start_time']:
        return False

    def to_time(val):
        if val is None: return None
        if isinstance(val, datetime): return val.time()
        import datetime as dt
        if isinstance(val, dt.timedelta):
            total = int(val.total_seconds())
            h, rem = divmod(total, 3600)
            m, s = divmod(rem, 60)
            return dt.time(h, m, s)
        return val

    start = to_time(cls['start_time'])
    if not start:
        return False

    now = datetime.now()
    start_dt = datetime.combine(now.date(), start)
    if now > start_dt + timedelta(minutes=tolerance_minutes):
        return True
    return False


def initialize_absent_records(class_id):
    cursor = get_db_cursor()

    # Optimized single INSERT using WHERE NOT EXISTS
    query = """
        INSERT INTO attendance (Student_id, class_id, date, status)
        SELECT s.Student_id, %s, CURDATE(), 'Absent'
        FROM student_data s
        WHERE s.Department_id = (
            SELECT Department_id FROM class_data WHERE class_id = %s
        )
        AND NOT EXISTS (
            SELECT 1 FROM attendance a 
            WHERE a.Student_id = s.Student_id 
              AND a.class_id = %s 
              AND a.date = CURDATE()
        )
    """
    cursor.execute(query, (class_id, class_id, class_id))
    conn.commit()
    affected = cursor.rowcount
    cursor.close()
    
    return affected
    
def mark_student_present(student_id, class_id, is_late=False):
    """
    Update the student's attendance record from 'Absent' → 'Present' or 'Late'.
    Uses UPDATE (not INSERT) because initialize_absent_records already
    created the row.  Returns True if a row was updated.
    """
    cursor = get_db_cursor()
    new_status = 'Late' if is_late else 'Present'
    cursor.execute(
        """
        UPDATE attendance
        SET status = %s
        WHERE Student_id = %s
          AND class_id   = %s
          AND date       = CURDATE()
          AND status     = 'Absent'
        """,
        (new_status, student_id, class_id)
    )
    conn.commit()
    affected = cursor.rowcount
    cursor.close()
    return affected > 0


# ----------------------------------------------------------------
#  API: Check if a class session is currently active
#  GET /api/session_status/<class_id>
# ----------------------------------------------------------------
@app.route("/api/session_status/<int:class_id>")
def api_session_status(class_id):
    if 'Student_id' not in session and 'faculty_id' not in session:
        return jsonify({"error": "Unauthorized"}), 401

    active = is_session_active(class_id)
    return jsonify({"class_id": class_id, "active": active})


# ----------------------------------------------------------------
#  API: Initialize absent records for a class (idempotent)
#  POST /api/init_attendance/<class_id>   (Faculty only)
# ----------------------------------------------------------------
@app.route("/api/init_attendance/<int:class_id>", methods=["POST"])
def api_init_attendance(class_id):
    if 'faculty_id' not in session:
        return jsonify({"error": "Unauthorized"}), 401

    if not is_session_active(class_id):
        return jsonify({
            "success": False,
            "message": "Class session is not currently active."
        }), 403

    inserted = initialize_absent_records(class_id)
    return jsonify({
        "success": True,
        "inserted": inserted,
        "message": f"{inserted} absent records created." if inserted else "Already initialized."
    })


# ----------------------------------------------------------------
#  API: Mark a student present (face-recognition or manual)
#  POST /api/mark_present
#  Body JSON: { "student_id": int, "class_id": int }
#  (Faculty / admin use — no face check)
# ----------------------------------------------------------------
@app.route("/api/mark_present", methods=["POST"])
def api_mark_present():
    if 'faculty_id' not in session and 'admin_id' not in session:
        return jsonify({"error": "Unauthorized"}), 401

    data       = request.get_json(force=True) or {}
    student_id = data.get("student_id")
    class_id   = data.get("class_id")

    if not student_id or not class_id:
        return jsonify({"success": False, "message": "Missing student_id or class_id"}), 400

    if not is_session_active(class_id):
        return jsonify({"success": False, "message": "Class session is not active."}), 403

    # Make sure the absent row exists first (in case init was skipped)
    initialize_absent_records(class_id)

    is_late = is_late_for_class(class_id)
    updated = mark_student_present(student_id, class_id, is_late=is_late)
    return jsonify({
        "success": updated,
        "message": "Marked Present." if updated else "No matching record found for today."
    })


# ----------------------------------------------------------------
#  Face-recognition attendance — updated to use UPDATE not INSERT
# ----------------------------------------------------------------
@app.route("/mark_attendanceprocess_v2", methods=["POST"])
def mark_attendanceprocess_v2():
    """
    New version of face-recognition attendance marking.
    Pre-seeds absent records, then updates to Present on match.
    """
    class_id   = request.form.get('class_id')
    student_id = session.get('Student_id')

    if not student_id:
        return redirect(url_for('student_login'))
    if not class_id:
        return jsonify({"success": False, "message": "Invalid class ID"}), 400

    class_id = int(class_id)

    # 1. Session active check
    if not is_session_active(class_id):
        return render_template(
            "Student/student_dashboard.html",
            msg="Class session is not currently active. Attendance cannot be marked."
        )

    # 2. Ensure absent row exists for this student today
    initialize_absent_records(class_id)

    cursor = get_db_cursor(dict_cursor=True)

    try:
        # 3. Decode webcam image
        image_data   = request.form['image_data'].split(",")[1]
        image_bytes  = base64.b64decode(image_data)
        np_arr       = np.frombuffer(image_bytes, np.uint8)
        frame        = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)
        rgb_frame    = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        rgb_frame    = cv2.resize(rgb_frame, (0, 0), fx=0.5, fy=0.5)

        unknown_encodings = face_recognition.face_encodings(rgb_frame)

        if len(unknown_encodings) == 0:
            return render_template("Student/student_dashboard.html",
                                   msg3="No face detected. Please try again.")
        if len(unknown_encodings) > 1:
            return render_template("Student/student_dashboard.html",
                                   msg4="Multiple faces detected. Please try alone.")

        unknown_encoding = unknown_encodings[0]

        # 4. Load stored student image
        cursor.execute(
            "SELECT Student_id, img_of_student FROM student_data WHERE Student_id=%s",
            (student_id,)
        )
        student = cursor.fetchone()
        if not student:
            return render_template("Student/student_dashboard.html",
                                   msg5="Student record not found.")

        img_path = student['img_of_student']
        if not img_path.startswith("static"):
            img_path = os.path.join("static/student_img_upload", img_path)

        if not os.path.exists(img_path):
            return render_template("Student/student_dashboard.html",
                                   msg6="Registration image missing. Contact admin.")

        known_image    = face_recognition.load_image_file(img_path)
        known_encodings = face_recognition.face_encodings(known_image)

        if len(known_encodings) == 0:
            return render_template("Student/student_dashboard.html",
                                   msg7="No face found in stored image.")

        known_encoding = known_encodings[0]

        # 5. Compare faces
        matches  = face_recognition.compare_faces([known_encoding], unknown_encoding, tolerance=0.55)
        distance = face_recognition.face_distance([known_encoding], unknown_encoding)
        print(f"[FaceRec] Match={matches[0]}, Distance={distance[0]:.4f}")

        if matches[0]:
            # 6. Check if already marked Present
            cursor.execute(
                "SELECT status FROM attendance "
                "WHERE Student_id=%s AND class_id=%s AND date=CURDATE()",
                (student['Student_id'], class_id)
            )
            existing = cursor.fetchone()

            if existing and existing['status'] == 'Present':
                return render_template("Student/student_dashboard.html",
                                       msg="Attendance already marked Present for today.")

            # 7. UPDATE to Present or Late
            is_late = is_late_for_class(class_id)
            updated = mark_student_present(student['Student_id'], class_id, is_late=is_late)
            if updated:
                msg1 = "✅ Late Attendance Marked Successfully!" if is_late else "✅ Attendance Marked Successfully!"
                return render_template("Student/student_dashboard.html",
                                       msg1=msg1)
            else:
                return render_template("Student/student_dashboard.html",
                                       msg="Could not update attendance. Try again.")
        else:
            return render_template("Student/student_dashboard.html",
                                   msg2="Face does not match. Attendance denied.")

    except Exception as e:
        print("[mark_attendanceprocess_v2] Error:", e)
        return render_template("Student/student_dashboard.html",
                               msg="An error occurred. Please try again.")
    finally:
        cursor.close()


# ----------------------------------------------------------------
#  Faculty view: live attendance list for a class
#  GET /faculty_class_attendance/<class_id>
# ----------------------------------------------------------------
@app.route("/faculty_class_attendance/<int:class_id>")
def faculty_class_attendance(class_id):
    if 'faculty_id' not in session:
        return redirect(url_for('faculty_login'))

    # Auto-initialize when faculty opens this page
    if is_session_active(class_id):
        initialize_absent_records(class_id)

    cursor = get_db_cursor(dict_cursor=True)

    # Fetch class info
    cursor.execute(
        """
        SELECT c.*, d.Department_name, f.Faculty_name
        FROM class_data c
        JOIN department_data d ON c.Department_id = d.Department_id
        JOIN faculty_data f    ON c.Faculty_id    = f.Faculty_id
        WHERE c.class_id = %s
        """,
        (class_id,)
    )
    cls = cursor.fetchone()

    # Fetch today's attendance for this class
    cursor.execute(
        """
        SELECT a.attendance_id, s.Student_id, s.Student_name,
               s.Enrollment_no, a.status
        FROM attendance a
        JOIN student_data s ON a.Student_id = s.Student_id
        WHERE a.class_id = %s AND a.date = CURDATE()
        ORDER BY s.Student_name
        """,
        (class_id,)
    )
    records = cursor.fetchall()
    cursor.close()

    session_active = is_session_active(class_id)

    return render_template(
        "Faculty/class_attendance.html",
        cls=cls,
        records=records,
        session_active=session_active,
        class_id=class_id
    )


if __name__ == '__main__':
    app.run(debug=True)

